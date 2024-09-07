from openpyxl import load_workbook
import re


convert_column = {
    "B": "W",
    "C": "E",
    "D": "B",
    "E": "G",
    "F": "P",
    "G": "R",
    "H": "Q",
    "J": "U",
}
break_col = "A"
pass_col = "C"

convert_vehicle = {
    "Thẻ ô tô tháng": "ÔTÔ THÁNG",
    "Thẻ Xe dap thang": "XE ĐẠP THÁNG",
    "Thẻ xe máy tháng": "XE MÁY THÁNG"
}

def convert_plate(plate: str) -> str:
    if re.match(r"^[0-9][0-9][aA-zZ]-", plate):
        pass
    elif re.match(r"^[0-9][0-9][aA-zZ]([0-9]|[aA-zZ])-", plate):
        pass
    elif re.match(r"^[0-9][0-9][aA-zZ]([0-9]|[aA-zZ])", plate):
        pass
    return plate

def convert_time(date_time: str) -> str:
    return date_time

convert_method = {
    "F": lambda v: convert_vehicle[v] if v in convert_vehicle else "",
    "G": convert_plate,
    "J": convert_time,
}

MAX_ROW = 10_000

out_temp = load_workbook("Mau dang ky thang.xlsx")
out_start = 5

input_temp = load_workbook("card.xlsx")
input_start = 2

out_sheet = out_temp.active
input_sheet = input_temp.active

ordinal = out_start
for i in range(input_start, MAX_ROW):
    if not input_sheet[break_col + str(i)].value:
        break
    if not input_sheet[pass_col + str(i)].value:
        continue
    for col in convert_column.keys():
        if col in convert_method:
            out_sheet[convert_column[col] + str(ordinal)].value = convert_method[col](input_sheet[col + str(i)].value)
        else:
            out_sheet[convert_column[col] + str(ordinal)].value = input_sheet[col + str(i)].value
    ordinal += 1

out_temp.save("output dang ki khach hang.xlsx")